import streamlit as st
import pandas as pd
import sqlite3
import json
import hashlib
from datetime import datetime, time
import os
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# تنظیمات اولیه صفحه Streamlit
st.set_page_config(
    page_title="سامانه مدیریت و بایگانی ساخت افزایشی (SLM)",
    page_icon="logo.png" if os.path.exists("logo.png") else "⚙️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =========================================================
# ۱. توابع عمومی، پردازشی و تبدیل تاریخ و زمان (Top-Level)
# =========================================================

def time_to_hours(t_val):
    if isinstance(t_val, time):
        return t_val.hour + (t_val.minute / 60.0)
    elif isinstance(t_val, str) and ":" in t_val:
        try:
            parts = t_val.strip().split(":")
            return int(parts[0]) + (int(parts[1]) / 60.0)
        except Exception:
            return 0.0
    try:
        return float(t_val)
    except Exception:
        return 0.0

def hours_to_time_str(hrs):
    try:
        hrs_float = float(hrs)
        h = int(hrs_float)
        m = int(round((hrs_float - h) * 60))
        if m == 60:
            h += 1
            m = 0
        return f"{h:02d}:{m:02d}"
    except Exception:
        return "00:00"

def gregorian_to_jalali(gy, gm, gd):
    g_d_m = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334]
    if gy > 1600:
        jy = 979
        gy -= 1600
    else:
        jy = 0
        gy -= 621
    gy2 = gy if gm > 2 else gy - 1
    days = (365 * gy) + ((gy2 + 4) // 4) - ((gy2 + 100) // 100) + ((gy2 + 400) // 400) - 80 + gd + g_d_m[gm - 1]
    jy += 33 * (days // 12053)
    days %= 12053
    jy += 4 * (days // 1461)
    days %= 1461
    if days > 365:
        jy += (days - 1) // 365
        days = (days - 1) % 365
    if days < 186:
        jm = 1 + (days // 31)
        jd = 1 + (days % 31)
    else:
        jm = 7 + ((days - 186) // 30)
        jd = 1 + ((days - 186) % 30)
    return f"{jy:04d}/{jm:02d}/{jd:02d}"

def get_today_jalali():
    now = datetime.now()
    return gregorian_to_jalali(now.year, now.month, now.day)

def format_jalali_date(date_str):
    if not date_str:
        return ""
    date_clean = str(date_str).strip()
    if "-" in date_clean:
        parts = date_clean.split("-")
        if len(parts) == 3 and len(parts[0]) == 4:
            try:
                return gregorian_to_jalali(int(parts[0]), int(parts[1]), int(parts[2]))
            except Exception:
                return date_clean
    return date_clean

def flatten_powder_df(df):
    if df.empty:
        return df
    records = []
    for _, row in df.iterrows():
        r_dict = {
            'شماره ظرف پودر': row.get('powder_code', ''),
            'جنس / نوع متریال پودر': row.get('material', ''),
            'وزن پودر (گرم)': row.get('weight_g', 0.0),
            'تاریخ ثبت (شمسی)': format_jalali_date(row.get('date', ''))
        }
        raw_json = row.get('checklist_json', '')
        if raw_json:
            try:
                chk = json.loads(raw_json)
                for test_name, test_val in chk.items():
                    if isinstance(test_val, dict):
                        r_dict[f"{test_name} (وضعیت)"] = test_val.get('status', '')
                        r_dict[f"{test_name} (ملاحظات)"] = test_val.get('note', '')
                    else:
                        r_dict[f"{test_name}"] = str(test_val)
            except Exception:
                pass
        records.append(r_dict)
    return pd.DataFrame(records)

def flatten_qc_df(df):
    if df.empty:
        return df
    records = []
    for _, row in df.iterrows():
        r_dict = {
            'کد/شناسه قطعه': row.get('part_code', ''),
            'نام قطعه': row.get('part_name', ''),
            'شماره ظرف پودر مصرفی': row.get('material', ''),
            'مدل دستگاه': row.get('machine_model', ''),
            'تاریخ ثبت بازرسی (شمسی)': format_jalali_date(row.get('date', '')),
            'بازرس کنترل کیفیت': row.get('qc_inspector', ''),
            'مسئول مهندسی کیفیت': row.get('qc_engineer', ''),
            'مدیر تضمین کیفیت': row.get('qa_manager', '')
        }
        raw_json = row.get('qc_checks_json', '')
        if raw_json:
            try:
                chk = json.loads(raw_json)
                for test_name, test_val in chk.items():
                    if isinstance(test_val, dict):
                        r_dict[f"{test_name} - نتیجه"] = test_val.get('result', '')
                        r_dict[f"{test_name} - ملاحظات"] = test_val.get('note', '')
                    else:
                        r_dict[f"{test_name}"] = str(test_val)
            except Exception:
                pass
        records.append(r_dict)
    return pd.DataFrame(records)

def format_production_df_view(df):
    if df.empty:
        return df
    df_copy = df.copy()
    date_cols = ['date', 'start_date', 'end_date', 'delivery_date']
    for col in date_cols:
        if col in df_copy.columns:
            df_copy[col] = df_copy[col].apply(format_jalali_date)
    
    time_cols = ['build_time_hrs', 'downtime_hrs', 'setup_time_hrs', 'cleaning_time_hrs']
    for col in time_cols:
        if col in df_copy.columns:
            df_copy[col] = df_copy[col].apply(hours_to_time_str)
            
    return df_copy

def format_cost_df_view(df):
    if df.empty:
        return df
    df_copy = df.copy()
    time_cols = ['print_time_hrs', 'design_time_hrs', 'post_process_time_hrs']
    for col in time_cols:
        if col in df_copy.columns:
            df_copy[col] = df_copy[col].apply(hours_to_time_str)
    return df_copy

# --- موتور پیشرفته تولید اکسل فرم‌محور و چندبرگه ---
def build_form_layout_excel(sections_dict, title="کاربرگ اختصاصی سامانه SLM"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]
    ws.sheet_view.rightToLeft = True

    title_font = Font(name='B Nazanin', size=15, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    
    sec_font = Font(name='B Nazanin', size=12, bold=True, color='FFFFFF')
    sec_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    
    hdr_font = Font(name='B Nazanin', size=11, bold=True, color='FFFFFF')
    hdr_fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')

    lbl_font = Font(name='B Nazanin', size=11, bold=True, color='1E293B')
    lbl_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    val_font = Font(name='B Nazanin', size=11, color='0F172A')
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.merge_cells('A1:F1')
    c_title = ws['A1']
    c_title.value = title
    c_title.font = title_font
    c_title.fill = title_fill
    c_title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    curr_row = 3

    for sec_title, sec_content in sections_dict.items():
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
        c_sec = ws.cell(row=curr_row, column=1, value=sec_title)
        c_sec.font = sec_font
        c_sec.fill = sec_fill
        c_sec.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[curr_row].height = 28
        curr_row += 1

        if isinstance(sec_content, list) and len(sec_content) > 0 and isinstance(sec_content[0], tuple):
            pairs = sec_content
            for i in range(0, len(pairs), 2):
                ws.row_dimensions[curr_row].height = 24
                l1, v1 = pairs[i]
                c_l1 = ws.cell(row=curr_row, column=1, value=str(l1))
                c_l1.font = lbl_font; c_l1.fill = lbl_fill; c_l1.alignment = Alignment(horizontal='right', vertical='center'); c_l1.border = thin_border

                ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=3)
                c_v1 = ws.cell(row=curr_row, column=2, value=str(v1))
                c_v1.font = val_font; c_v1.alignment = Alignment(horizontal='right', vertical='center'); c_v1.border = thin_border
                ws.cell(row=curr_row, column=3).border = thin_border

                if i + 1 < len(pairs):
                    l2, v2 = pairs[i+1]
                    c_l2 = ws.cell(row=curr_row, column=4, value=str(l2))
                    c_l2.font = lbl_font; c_l2.fill = lbl_fill; c_l2.alignment = Alignment(horizontal='right', vertical='center'); c_l2.border = thin_border

                    ws.merge_cells(start_row=curr_row, start_column=5, end_row=curr_row, end_column=6)
                    c_v2 = ws.cell(row=curr_row, column=5, value=str(v2))
                    c_v2.font = val_font; c_v2.alignment = Alignment(horizontal='right', vertical='center'); c_v2.border = thin_border
                    ws.cell(row=curr_row, column=6).border = thin_border
                else:
                    ws.merge_cells(start_row=curr_row, start_column=4, end_row=curr_row, end_column=6)

                curr_row += 1

        elif isinstance(sec_content, dict):
            ws.row_dimensions[curr_row].height = 24
            
            c_h1 = ws.cell(row=curr_row, column=1, value="ردیف")
            c_h1.font = hdr_font; c_h1.fill = hdr_fill; c_h1.alignment = Alignment(horizontal='center', vertical='center'); c_h1.border = thin_border
            
            ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=3)
            c_h2 = ws.cell(row=curr_row, column=2, value="عنوان آزمون / پارامتر کیفی")
            c_h2.font = hdr_font; c_h2.fill = hdr_fill; c_h2.alignment = Alignment(horizontal='right', vertical='center'); c_h2.border = thin_border
            ws.cell(row=curr_row, column=3).border = thin_border
            
            c_h3 = ws.cell(row=curr_row, column=4, value="وضعیت / نتیجه")
            c_h3.font = hdr_font; c_h3.fill = hdr_fill; c_h3.alignment = Alignment(horizontal='center', vertical='center'); c_h3.border = thin_border
            
            ws.merge_cells(start_row=curr_row, start_column=5, end_row=curr_row, end_column=6)
            c_h4 = ws.cell(row=curr_row, column=5, value="ملاحظات و توضیحات")
            c_h4.font = hdr_font; c_h4.fill = hdr_fill; c_h4.alignment = Alignment(horizontal='right', vertical='center'); c_h4.border = thin_border
            ws.cell(row=curr_row, column=6).border = thin_border

            curr_row += 1
            
            idx = 1
            for test_name, test_data in sec_content.items():
                ws.row_dimensions[curr_row].height = 22
                status_val = test_data.get('status', test_data.get('result', '')) if isinstance(test_data, dict) else str(test_data)
                note_val = test_data.get('note', '') if isinstance(test_data, dict) else ''

                c_idx = ws.cell(row=curr_row, column=1, value=idx)
                c_idx.font = val_font; c_idx.alignment = Alignment(horizontal='center', vertical='center'); c_idx.border = thin_border
                
                ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=3)
                c_tn = ws.cell(row=curr_row, column=2, value=test_name)
                c_tn.font = val_font; c_tn.alignment = Alignment(horizontal='right', vertical='center'); c_tn.border = thin_border
                ws.cell(row=curr_row, column=3).border = thin_border
                
                c_st = ws.cell(row=curr_row, column=4, value=status_val)
                c_st.font = val_font; c_st.alignment = Alignment(horizontal='center', vertical='center'); c_st.border = thin_border
                
                ws.merge_cells(start_row=curr_row, start_column=5, end_row=curr_row, end_column=6)
                c_nt = ws.cell(row=curr_row, column=5, value=note_val)
                c_nt.font = val_font; c_nt.alignment = Alignment(horizontal='right', vertical='center'); c_nt.border = thin_border
                ws.cell(row=curr_row, column=6).border = thin_border

                curr_row += 1
                idx += 1

        curr_row += 1

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_to_styled_excel_multisheet(dict_of_dfs, file_name="export.xlsx"):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in dict_of_dfs.items():
            if df.empty:
                df = pd.DataFrame(["اطلاعاتی ثبت نشده است"], columns=["وضعیت"])
            
            clean_sheet_name = sheet_name[:30].replace(":", "").replace("?", "").replace("*", "").replace("/", "").replace("\\", "")
            df.to_excel(writer, index=False, sheet_name=clean_sheet_name)
            worksheet = writer.sheets[clean_sheet_name]

            worksheet.sheet_view.rightToLeft = True

            header_font = Font(name='B Nazanin', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
            data_font = Font(name='B Nazanin', size=11)
            align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )

            for col_num, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = data_font
                    cell.alignment = align_right
                    cell.border = thin_border

            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max(max_len + 5, 16)

    output.seek(0)
    return output

# --- استایل تمیز و راست‌چین CSS ---
st.markdown("""
<style>
    @font-face {
        font-family: 'B Nazanin';
        src: url('https://cdn.fontcdn.ir/Font/Persian/BNazanin/BNazanin.eot');
        src: url('https://cdn.fontcdn.ir/Font/Persian/BNazanin/BNazanin.eot?#iefix') format('embedded-opentype'),
             url('https://cdn.fontcdn.ir/Font/Persian/BNazanin/BNazanin.woff') format('woff'),
             url('https://cdn.fontcdn.ir/Font/Persian/BNazanin/BNazanin.ttf') format('truetype');
        font-weight: normal;
        font-style: normal;
    }

    [data-testid="stInputInstruction"],
    div[data-testid="stInputInstruction"],
    .stInputInstruction,
    div:has(> input) ~ div,
    div:has(> textarea) ~ div,
    [aria-live="polite"],
    small {
        display: none !important;
        visibility: hidden !important;
        opacity: 0 !important;
        height: 0px !important;
        width: 0px !important;
        margin: 0px !important;
        padding: 0px !important;
        pointer-events: none !important;
    }

    [data-testid="stHeaderActionElements"],
    .header-anchor,
    a[href^="#"] {
        display: none !important;
    }

    [data-testid="stMainBlockContainer"],
    [data-testid="stSidebarUserContent"] {
        direction: rtl !important;
        text-align: right !important;
        font-family: 'B Nazanin', 'Vazir', sans-serif !important;
    }

    h1, h2, h3, h4, h5, h6, p, label, button, table {
        font-family: 'B Nazanin', 'Vazir', sans-serif !important;
        direction: rtl !important;
        text-align: right !important;
    }

    p, span, label, button {
        font-size: 1.2rem !important;
    }

    h1 { font-size: 2.2rem !important; font-weight: bold; }
    h2 { font-size: 1.8rem !important; font-weight: bold; }
    h3 { font-size: 1.5rem !important; font-weight: bold; }

    input, select, textarea, div[data-baseweb="input"] input, div[data-baseweb="select"] div {
        font-family: 'B Nazanin', 'Vazir', sans-serif !important;
        direction: rtl !important;
        text-align: right !important;
        font-size: 1.2rem !important;
    }

    table, [data-testid="stTable"], .stTable {
        direction: rtl !important;
        width: 100% !important;
        font-size: 1.15rem !important;
        border-collapse: collapse !important;
        text-align: right !important;
    }
    
    th, td, table th, table td {
        text-align: right !important;
        direction: rtl !important;
        padding: 10px !important;
    }

    div[data-testid="stSidebar"] [data-testid="stRadio"] > div {
        display: flex;
        flex-direction: column;
        gap: 8px;
    }

    div[data-testid="stSidebar"] [data-testid="stRadio"] label {
        background-color: transparent !important;
        border-radius: 8px !important;
        padding: 8px 16px !important;
        cursor: pointer !important;
        transition: all 0.2s ease !important;
        border: 1px solid transparent !important;
        direction: rtl !important;
        text-align: right !important;
    }

    div[data-testid="stSidebar"] [data-testid="stRadio"] label:hover {
        background-color: #2a374e !important;
        border: 1px solid #3b82f6 !important;
    }

    div[data-testid="stSidebar"] [data-testid="stRadio"] label[data-checked="true"] {
        background-color: #334155 !important;
        border: 1px solid #475569 !important;
        box-shadow: 0 2px 4px rgba(0,0,0,0.3) !important;
    }

    div[data-testid="stSidebar"] [data-testid="stRadio"] input[type="radio"] {
        display: none !important;
    }

    [data-testid="stMetric"] {
        background-color: #1e293b !important;
        border: 1px solid #334155 !important;
        border-radius: 10px !important;
        padding: 15px !important;
        text-align: center !important;
    }
    [data-testid="stMetricLabel"] > div {
        color: #94a3b8 !important;
        font-size: 1.2rem !important;
        justify-content: center !important;
    }
    [data-testid="stMetricValue"] > div {
        color: #38bdf8 !important;
        font-size: 2.2rem !important;
        justify-content: center !important;
    }

    .stButton>button, .stDownloadButton>button {
        width: 100%;
        background-color: #2563eb !important;
        color: #ffffff !important;
        border-radius: 8px !important;
        font-weight: bold !important;
        font-size: 1.2rem !important;
        padding: 10px 20px !important;
        border: none !important;
    }
</style>
""", unsafe_allow_html=True)

FARSI_HEADERS_MAP = {
    'powder_code': 'شماره ظرف پودر',
    'material': 'جنس / نوع متریال پودر',
    'weight_g': 'وزن پودر (گرم)',
    'date': 'تاریخ ثبت (شمسی)',
    'id': 'شناسه',
    'recycled_batch_code': 'کد پارت بازیافت',
    'input_powder_g': 'پودر ورودی به دستگاه (گرم)',
    'unrecyclable_powder_g': 'پودر غیرقابل بازیافت (گرم)',
    'recycled_powder_g': 'پودر بازیافتی قابل استفاده (گرم)',
    'notes': 'توضیحات و ملاحظات',
    'part_code': 'کد/شناسه قطعه',
    'part_name': 'نام قطعه',
    'quantity': 'تعداد روی صفحه',
    'machine_model': 'مدل دستگاه',
    'build_time_hrs': 'زمان تولید (ساعت:دقیقه)',
    'downtime_hrs': 'زمان توقف (ساعت:دقیقه)',
    'start_date': 'تاریخ شروع (شمسی)',
    'start_time': 'ساعت شروع',
    'end_date': 'تاریخ پایان (شمسی)',
    'end_time': 'ساعت پایان',
    'setup_time_hrs': 'زمان آماده‌سازی (ساعت:دقیقه)',
    'cleaning_time_hrs': 'زمان تمیزکاری (ساعت:دقیقه)',
    'waste_powder_g': 'پودر غیرقابل بازیافت (گرم)',
    'part_with_support_g': 'وزن با ساپورت (گرم)',
    'final_part_g': 'وزن قطعه نهایی (گرم)',
    'filter_percentage': 'درصد فیلتر دستگاه (%)',
    'build_plate_code': 'کد صفحه ساخت',
    'build_plate_init_wt_g': 'وزن اولیه صفحه ساخت (گرم)',
    'build_plate_post_wt_g': 'وزن صفحه ساخت بعد پرداخت (گرم)',
    'engraving_qty': 'تعداد حکاکی لیزر',
    'delivery_date': 'تاریخ تحویل (شمسی)',
    'qc_inspector': 'بازرس کنترل کیفیت',
    'qc_engineer': 'مسئول مهندسی کیفیت',
    'qa_manager': 'مدیر تضمین کیفیت',
    'powder_type': 'نوع پودر فلزی',
    'volume_cm3': 'حجم قطعه (cm3)',
    'net_weight_g': 'وزن خالص (گرم)',
    'support_volume_cm3': 'حجم ساپورت (cm3)',
    'support_weight_g': 'وزن ساپورت (گرم)',
    'machine_type': 'نوع دستگاه',
    'parts_on_plate': 'تعداد قطعات روی صفحه',
    'print_time_hrs': 'زمان چاپ (ساعت:دقیقه)',
    'design_time_hrs': 'زمان طراحی (ساعت:دقیقه)',
    'post_process_time_hrs': 'زمان پرداخت (ساعت:دقیقه)',
    'overhead_pct': 'ضریب سربار (%)',
    'powder_cost_total': 'هزینه پودر (ریال)',
    'argon_cost_total': 'هزینه گاز آرگون (ریال)',
    'depreciation_cost_total': 'هزینه استهلاک دستگاه (ریال)',
    'power_cost_total': 'هزینه برق (ریال)',
    'engineering_cost_total': 'هزینه طراحی/مهندسی (ریال)',
    'operator_cost_total': 'هزینه اپراتور (ریال)',
    'post_process_cost_total': 'هزینه پرداخت‌کاری (ریال)',
    'qc_cost_total': 'هزینه کنترل کیفیت (ریال)',
    'utility_ventilation': 'هزینه تهویه (ریال)',
    'utility_chiller': 'هزینه چیلر (ریال)',
    'total_production_cost': 'بهای تمام شده کل (ریال)',
    'overhead_cost': 'مبلغ سربار (ریال)',
    'final_price': 'قیمت نهایی قابل ارائه به مشتری (ریال)'
}

DB_NAME = "slm_management.db"

DEFAULT_RATES = {
    "powder_price_Steel_316": 120000000,
    "powder_price_Ti6Al4V": 500000000,
    "powder_price_Inconel_718": 300000000,
    "powder_price_Hastelloy_X": 400000000,
    "machine_depr_M120": 1250000,
    "machine_depr_M300": 3125000,
    "power_kw_M120": 7,
    "power_kw_M300": 15,
    "wage_designer": 2500000,
    "wage_operator": 1870000,
    "wage_qc": 2180000,
    "argon_rate": 300000,
    "electricity_rate": 50000,
    "post_process_rate": 500000,
    "qc_fixed_cost": 40000000,
    "ventilation_rate": 200000,
    "chiller_rate": 2500
}

def hash_password(password):
    return hashlib.sha256(str.encode(password)).hexdigest()

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (username TEXT PRIMARY KEY, password_hash TEXT, full_name TEXT, role TEXT)''')
    
    default_users = [
        ('admin', hash_password('admin123'), 'مدیر سیستم', 'مدیریت'),
        ('operator', hash_password('op123'), 'اپراتور و طراح', 'اپراتور و طراح'),
        ('qc_user', hash_password('qc123'), 'کارشناس QC', 'کنترل کیفیت'),
        ('commerce', hash_password('com123'), 'کارشناس بازرگانی', 'بازرگانی')
    ]
    for u, p, f, r in default_users:
        c.execute("INSERT OR IGNORE INTO users (username, password_hash, full_name, role) VALUES (?, ?, ?, ?)", (u, p, f, r))

    c.execute('''CREATE TABLE IF NOT EXISTS system_settings (key TEXT PRIMARY KEY, value REAL)''')
    for key, val in DEFAULT_RATES.items():
        c.execute("INSERT OR IGNORE INTO system_settings (key, value) VALUES (?, ?)", (key, val))

    c.execute('''CREATE TABLE IF NOT EXISTS powders (powder_code TEXT PRIMARY KEY, material TEXT, weight_g REAL, date TEXT, checklist_json TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS recycled_powders (id INTEGER PRIMARY KEY AUTOINCREMENT, powder_code TEXT, recycled_batch_code TEXT, input_powder_g REAL, unrecyclable_powder_g REAL, recycled_powder_g REAL, date TEXT, notes TEXT, FOREIGN KEY(powder_code) REFERENCES powders(powder_code))''')
    c.execute('''CREATE TABLE IF NOT EXISTS nora_powders (powder_code TEXT PRIMARY KEY, material TEXT, weight_g REAL, date TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS production (part_code TEXT PRIMARY KEY, part_name TEXT, powder_code TEXT, quantity INTEGER, machine_model TEXT, date TEXT, build_time_hrs REAL, downtime_hrs REAL, start_date TEXT, start_time TEXT, end_date TEXT, end_time TEXT, setup_time_hrs REAL, cleaning_time_hrs REAL, input_powder_g REAL, waste_powder_g REAL, part_with_support_g REAL, final_part_g REAL, filter_percentage REAL, build_plate_code TEXT, build_plate_init_wt_g REAL, build_plate_post_wt_g REAL, finishing_json TEXT, engraving_qty INTEGER, delivery_date TEXT, FOREIGN KEY(powder_code) REFERENCES powders(powder_code))''')
    c.execute('''CREATE TABLE IF NOT EXISTS qc (part_code TEXT PRIMARY KEY, part_name TEXT, material TEXT, machine_model TEXT, date TEXT, qc_checks_json TEXT, qc_inspector TEXT, qc_engineer TEXT, qa_manager TEXT, FOREIGN KEY(part_code) REFERENCES production(part_code))''')
    c.execute('''CREATE TABLE IF NOT EXISTS cost_calculator (part_code TEXT PRIMARY KEY, part_name TEXT, powder_type TEXT, volume_cm3 REAL, net_weight_g REAL, support_volume_cm3 REAL, support_weight_g REAL, machine_type TEXT, parts_on_plate INTEGER, print_time_hrs REAL, design_time_hrs REAL, post_process_time_hrs REAL, overhead_pct REAL, powder_cost_total REAL, argon_cost_total REAL, depreciation_cost_total REAL, power_cost_total REAL, engineering_cost_total REAL, operator_cost_total REAL, post_process_cost_total REAL, qc_cost_total REAL, utility_ventilation REAL, utility_chiller REAL, total_production_cost REAL, overhead_cost REAL, final_price REAL, FOREIGN KEY(part_code) REFERENCES production(part_code))''')
    
    conn.commit()
    conn.close()

init_db()

def get_db_connection():
    return sqlite3.connect(DB_NAME)

def load_system_rates():
    conn = get_db_connection()
    df = pd.read_sql_query("SELECT * FROM system_settings", conn)
    conn.close()
    rates = DEFAULT_RATES.copy()
    for _, row in df.iterrows():
        rates[row['key']] = row['value']
    return rates

if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
    st.session_state["username"] = ""
    st.session_state["user_role"] = ""
    st.session_state["full_name"] = ""

def login_user(username, password):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT full_name, role FROM users WHERE username=? AND password_hash=?", (username, hash_password(password)))
    result = c.fetchone()
    conn.close()
    if result:
        st.session_state["authenticated"] = True
        st.session_state["username"] = username
        st.session_state["full_name"] = result[0]
        st.session_state["user_role"] = result[1]
        return True
    return False

def logout_user():
    st.session_state["authenticated"] = False
    st.session_state["username"] = ""
    st.session_state["user_role"] = ""
    st.session_state["full_name"] = ""
    st.rerun()

# ---------------------------------------------------------
# صفحه ورود
# ---------------------------------------------------------
if not st.session_state["authenticated"]:
    col_login_1, col_login_2, col_login_3 = st.columns([1, 2, 1])
    with col_login_2:
        st.markdown("<br>", unsafe_allow_html=True)
        if os.path.exists("logo.png"):
            st.image("logo.png", width=160)
        
        st.title("🔐 ورود به سامانه ساخت افزایشی (SLM)")
        st.caption("لطفاً نام کاربری و رمز عبور خود را وارد کنید:")
        
        with st.form("login_form"):
            username_input = st.text_input("نام کاربری (Username)")
            password_input = st.text_input("رمز عبور (Password)", type="password")
            submit_login = st.form_submit_button("🔑 ورود به سامانه")
            
            if submit_login:
                if login_user(username_input, password_input):
                    st.success(f"خوش آمدید {st.session_state['full_name']}!")
                    st.rerun()
                else:
                    st.error("نام کاربری یا رمز عبور اشتباه است.")

# ---------------------------------------------------------
# برنامه اصلی پس از ورود
# ---------------------------------------------------------
else:
    role = st.session_state["user_role"]
    
    if os.path.exists("logo.png"):
        st.sidebar.image("logo.png", width=140)
        
    st.sidebar.markdown(f"👤 **کاربر:** {st.session_state['full_name']}")
    st.sidebar.markdown(f"مقام/سطح: `{role}`")
    if st.sidebar.button("🚪 خروج از حساب"):
        logout_user()
    st.sidebar.markdown("---")

    if role == "مدیریت":
        menu_options = ["🏠 خانه", "📦 پودر", "🏭 فرم تولید", "❇️ کنترل کیفیت", "💰 محاسبه‌گر هزینه", "⚙️ تنظیمات نرخ‌ها", "🔍 بایگانی", "👥 مدیریت کاربران"]
    elif role == "اپراتور و طراح":
        menu_options = ["🏠 خانه", "📦 پودر", "🏭 فرم تولید"]
    elif role == "کنترل کیفیت":
        menu_options = ["🏠 خانه", "❇️ کنترل کیفیت", "🔍 بایگانی"]
    elif role == "بازرگانی":
        menu_options = ["🏠 خانه", "📦 پودر", "💰 محاسبه‌گر هزینه", "🔍 بایگانی"]
    else:
        menu_options = ["🏠 خانه"]

    choice = st.sidebar.radio("", menu_options, index=0)

    # ---------------------------------------------------------
    # ۰. خانه
    # ---------------------------------------------------------
    if choice == "🏠 خانه":
        head_c1, head_col2 = st.columns([6, 1])
        with head_c1:
            st.title("🧩 سامانه بایگانی و مدیریت تولید قطعات چاپ سه بعدی (SLM)")
            st.caption("این سامانه فرم‌های «پودر»، «فرم تولید»، «کنترل کیفیت (QC)» و «محاسبه‌گر قیمت» را به هم مرتبط کرده و امکان بایگانی را فراهم می‌سازد.")
        with head_col2:
            if os.path.exists("logo.png"):
                st.image("logo.png", width=120)

        st.markdown("<br>", unsafe_allow_html=True)
        
        conn = get_db_connection()
        powder_count = pd.read_sql_query("SELECT COUNT(*) as cnt FROM powders", conn)['cnt'].values[0] + pd.read_sql_query("SELECT COUNT(*) as cnt FROM nora_powders", conn)['cnt'].values[0]
        prod_count = pd.read_sql_query("SELECT COUNT(*) as cnt FROM production", conn)['cnt'].values[0]
        qc_count = pd.read_sql_query("SELECT COUNT(*) as cnt FROM qc", conn)['cnt'].values[0]
        cost_count = pd.read_sql_query("SELECT COUNT(*) as cnt FROM cost_calculator", conn)['cnt'].values[0]
        conn.close()

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("رکوردهای پودر", powder_count)
        m2.metric("فرم‌های تولید", prod_count)
        m3.metric("فرم‌های کنترل کیفیت", qc_count)
        m4.metric("محاسبات هزینه", cost_count)

        st.markdown("---")
        st.subheader("راهنمای گردش کار")
        st.markdown(f"""
        1. **📦 پودر** — ثبت و ویرایش ظروف پودر اولیه، بازیافتی و نورا با خروجی اکسل کامل چک‌لیست و یادداشت‌ها.
        2. **🏭 فرم تولید** — ثبت و ویرایش پارامترهای ساخت، زمان‌ها (HH:MM) و اوزان با خروجی اکسل فرم‌محور.
        3. **❇️ کنترل کیفیت (QC)** — ثبت و ویرایش نتایج تست‌ها و نوت‌ها با فرمت اکسل کاربرگی رسمی.
        4. **💰 محاسبه‌گر هزینه** — برآورد خودکار بهای تمام شده و قیمت فروش با آخرین نرخ‌های روز.
        5. **🔍 بایگانی و جستجو** — استعلام شناسنامه جامع قطعات و دریافت کاربرگ چندبرگه کامل.
        
        <br>
        <p style='color: #94a3b8; font-size: 1.1rem;'>امروز: {get_today_jalali()} | از منوی سمت راست بین صفحات جابه‌جا شوید.</p>
        """, unsafe_allow_html=True)

    # ---------------------------------------------------------
    # ۱. پودر
    # ---------------------------------------------------------
    elif choice == "📦 پودر":
        st.header("🧪 فرم مدیریت، آنالیز و بایگانی پودر")
        
        conn = get_db_connection()
        powders_df = pd.read_sql_query("SELECT * FROM powders", conn)
        nora_df = pd.read_sql_query("SELECT * FROM nora_powders", conn)
        recycled_df = pd.read_sql_query("SELECT * FROM recycled_powders", conn)
        
        sub_tab1, sub_tab2, sub_tab3 = st.tabs(["📦 ۱- پودرهای خریداری شده اولیه", "♻️ ۲- پودرهای بازیافت شده", "🏭 ۳- پودرهای خریداری شده از نورا"])
        
        with sub_tab1:
            st.subheader("حالت ثبت / ویرایش پودر اولیه")
            mode_pwd_init = st.radio("", ["ثبت پودر جدید", "ویرایش پودر موجود"], horizontal=True, key="mode_pwd_init")
            
            edit_pwd_init_data = {}
            chk_init_saved = {}
            if mode_pwd_init == "ویرایش پودر موجود":
                existing_pwds = powders_df['powder_code'].tolist()
                if existing_pwds:
                    sel_pwd_code = st.selectbox("انتخاب کد پودر جهت ویرایش:", existing_pwds, key="sel_pwd_init_edit")
                    edit_pwd_init_data = powders_df[powders_df['powder_code'] == sel_pwd_code].iloc[0].to_dict()
                    raw_chk = edit_pwd_init_data.get('checklist_json', '')
                    if raw_chk:
                        try:
                            chk_init_saved = json.loads(raw_chk)
                        except Exception:
                            pass
                else:
                    st.warning("هنوز هیچ پودر اولیه‌ای ثبت نشده است.")

            with st.expander("📝 فرم اطلاعات و چک‌لیست پودر اولیه", expanded=True):
                col1, col2 = st.columns(2)
                with col1:
                    def_pwd_code = edit_pwd_init_data.get('powder_code', '') if mode_pwd_init == "ویرایش پودر موجود" else ""
                    powder_code = st.text_input("کد/شماره ظرف پودر فلز (مانند: PWD-316-01)", value=def_pwd_code, disabled=(mode_pwd_init == "ویرایش پودر موجود"))
                    
                    mat_options = ["Steel 316", "Ti6Al4V", "Inconel 718", "Hastelloy X", "سایر"]
                    def_mat_idx = mat_options.index(edit_pwd_init_data['material']) if mode_pwd_init == "ویرایش پودر موجود" and edit_pwd_init_data.get('material') in mat_options else 0
                    material = st.selectbox("جنس پودر", mat_options, index=def_mat_idx)
                with col2:
                    def_wt = float(edit_pwd_init_data.get('weight_g', 10000.0))
                    weight_g = st.number_input("وزن پودر (گرم)", min_value=0.0, value=def_wt)
                    
                    def_date_val = format_jalali_date(edit_pwd_init_data.get('date', '')) if mode_pwd_init == "ویرایش پودر موجود" else get_today_jalali()
                    date_str = st.text_input("تاریخ ورود/تست (هجری شمسی)", value=def_date_val)
                    
                st.markdown("---")
                st.subheader("📋 چک‌لیست آزمون‌های خواص پودر")
                checklist = [
                    "بررسی خلوص شیمیایی", "اندازه‌گیری میزان رطوبت", "اندازه‌گیری چگالی ضربه‌ای",
                    "اندازه‌گیری چگالی ظاهری", "بررسی مورفولوژی ذرات", "آنالیز توزیع اندازه ذرات",
                    "وضعیت بسته‌بندی و عدم آلودگی", "بررسی گواهینامه کیفیت تامین‌کننده"
                ]
                
                qc_results = {}
                for item in checklist:
                    c1, c2, c3 = st.columns([3, 2, 4])
                    saved_item = chk_init_saved.get(item, {})
                    def_status = saved_item.get('status', 'تایید') if isinstance(saved_item, dict) else 'تایید'
                    def_note = saved_item.get('note', '') if isinstance(saved_item, dict) else ''
                    
                    with c1: st.write(f"**{item}**")
                    with c2: status = st.selectbox("وضعیت", ["تایید", "رد"], index=0 if def_status=="تایید" else 1, key=f"chk_p_{item}")
                    with c3: note = st.text_input("ملاحظات / توضیحات", value=def_note, key=f"note_p_{item}")
                    qc_results[item] = {"status": status, "note": note}
                    
                btn_c1, btn_c2 = st.columns([3, 2])
                with btn_c1:
                    save_pwd_init = st.button("💾 ذخیره در بایگانی پودر اولیه")
                    if save_pwd_init:
                        target_pwd_code = def_pwd_code if mode_pwd_init == "ویرایش پودر موجود" else powder_code
                        if target_pwd_code:
                            c = conn.cursor()
                            c.execute("""INSERT OR REPLACE INTO powders 
                                         (powder_code, material, weight_g, date, checklist_json)
                                         VALUES (?, ?, ?, ?, ?)""",
                                      (target_pwd_code, material, weight_g, date_str, json.dumps(qc_results, ensure_ascii=False)))
                            conn.commit()
                            st.success(f"اطلاعات پودر {target_pwd_code} با موفقیت ذخیره شد.")
                            st.rerun()
                        else:
                            st.error("لطفاً کد/شماره ظرف پودر را وارد کنید.")
                with btn_c2:
                    target_pwd_for_export = def_pwd_code if mode_pwd_init == "ویرایش پودر موجود" else powder_code
                    export_sec = {
                        "مشخصات عمومی ظرف پودر اولیه": [
                            ("کد / شماره ظرف پودر", target_pwd_for_export or "---"),
                            ("جنس / متریال پودر", material),
                            ("وزن پودر (گرم)", f"{weight_g:,.1f}"),
                            ("تاریخ ورود / تست (شمسی)", date_str)
                        ],
                        "چک‌لیست نتایج آزمون‌های خواص پودر و ملاحظات": qc_results
                    }
                    excel_form_bytes = build_form_layout_excel(export_sec, title=f"شناسنامه آزمون پودر - {target_pwd_for_export or 'جدید'}")
                    st.download_button(
                        label="📥 دانلود خروجی اکسل",
                        data=excel_form_bytes,
                        file_name=f"powder_form_{target_pwd_for_export or 'record'}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="btn_down_pwd_form_styled"
                    )

        with sub_tab2:
            st.subheader("♻️ مدیریت و ثبت پودرهای بازیافت شده")
            p_initial_list = [f"{code} [اولیه]" for code in powders_df['powder_code'].tolist()] if not powders_df.empty else []
            p_nora_list = [f"{code} [نورا]" for code in nora_df['powder_code'].tolist()] if not nora_df.empty else []
            combined_source_powders = p_initial_list + p_nora_list
            
            if not combined_source_powders:
                st.warning("ابتدا باید حداقل یک ظرف پودر اولیه یا پودر نورا ثبت کرده باشید.")
            else:
                st.subheader("حالت ثبت / ویرایش پودر بازیافتی")
                mode_rec = st.radio("", ["ثبت پارت جدید", "ویرایش پارت موجود"], horizontal=True, key="mode_rec")
                
                edit_rec_data = {}
                if mode_rec == "ویرایش پارت موجود":
                    if not recycled_df.empty:
                        rec_batches = recycled_df['recycled_batch_code'].tolist()
                        sel_rec_batch = st.selectbox("انتخاب شناسه پارت بازیافت جهت ویرایش:", rec_batches, key="sel_rec_batch")
                        edit_rec_data = recycled_df[recycled_df['recycled_batch_code'] == sel_rec_batch].iloc[0].to_dict()
                    else:
                        st.warning("هیچ رکوردی برای بازیافت ثبت نشده است.")

                rc1, rc2 = st.columns(2)
                with rc1:
                    saved_src = edit_rec_data.get('powder_code', '')
                    def_src_idx = 0
                    for idx, s_name in enumerate(combined_source_powders):
                        if s_name.startswith(saved_src):
                            def_src_idx = idx
                            break
                    selected_source = st.selectbox("شماره/کد ظرف پودر مبدا (خریداری‌شده اولیه / نورا)", combined_source_powders, index=def_src_idx)
                    clean_source_code = selected_source.split(" [")[0] if selected_source else ""
                    
                    def_rec_batch = edit_rec_data.get('recycled_batch_code', '') if mode_rec == "ویرایش پارت موجود" else ""
                    recycled_batch_code = st.text_input("شناسه پارت بازیافت (مانند: REC-PWD-01)", value=def_rec_batch)
                    
                    def_inp_wt = float(edit_rec_data.get('input_powder_g', 5000.0))
                    input_powder_g = st.number_input("پودر ورودی به دستگاه (گرم)", min_value=0.0, value=def_inp_wt)
                with rc2:
                    def_unrec_wt = float(edit_rec_data.get('unrecyclable_powder_g', 200.0))
                    unrecyclable_powder_g = st.number_input("پودر مصرف شده غیر قابل بازیافت (گرم)", min_value=0.0, value=def_unrec_wt)
                    recycled_powder_g = input_powder_g - unrecyclable_powder_g
                    st.metric("مقدار پودر بازیافت‌شده قابل استفاده (گرم)", f"{recycled_powder_g:,.1f}")
                    
                    def_rec_dt = format_jalali_date(edit_rec_data.get('date', '')) if mode_rec == "ویرایش پارت موجود" else get_today_jalali()
                    rec_date = st.text_input("تاریخ بازیافت (هجری شمسی)", value=def_rec_dt)
                    rec_notes = st.text_input("توضیحات و ملاحظات غربال‌گری / الک", value=edit_rec_data.get('notes', ''))
                
                r_btn1, r_btn2 = st.columns([3, 2])
                with r_btn1:
                    rec_submit = st.button("💾 ذخیره رکورد پودر بازیافتی")
                    if rec_submit:
                        if recycled_batch_code:
                            c = conn.cursor()
                            if mode_rec == "ویرایش پارت موجود" and 'id' in edit_rec_data:
                                c.execute("""UPDATE recycled_powders SET
                                             powder_code=?, recycled_batch_code=?, input_powder_g=?, unrecyclable_powder_g=?, recycled_powder_g=?, date=?, notes=?
                                             WHERE id=?""",
                                          (clean_source_code, recycled_batch_code, input_powder_g, unrecyclable_powder_g, recycled_powder_g, rec_date, rec_notes, edit_rec_data['id']))
                            else:
                                c.execute("""INSERT INTO recycled_powders 
                                             (powder_code, recycled_batch_code, input_powder_g, unrecyclable_powder_g, recycled_powder_g, date, notes)
                                             VALUES (?, ?, ?, ?, ?, ?, ?)""",
                                          (clean_source_code, recycled_batch_code, input_powder_g, unrecyclable_powder_g, recycled_powder_g, rec_date, rec_notes))
                            conn.commit()
                            st.success("اطلاعات پودر بازیافتی با موفقیت ثبت شد.")
                            st.rerun()
                        else:
                            st.error("شناسه پارت بازیافت الزامی است.")
                with r_btn2:
                    export_rec_sec = {
                        "مشخصات پارت پودر بازیافتی": [
                            ("کد ظرف پودر مبدا", clean_source_code),
                            ("شناسه پارت بازیافت", recycled_batch_code or "---"),
                            ("پودر ورودی به دستگاه (گرم)", f"{input_powder_g:,.1f}"),
                            ("پودر غیرقابل بازیافت (گرم)", f"{unrecyclable_powder_g:,.1f}"),
                            ("پودر بازیافتی قابل استفاده (گرم)", f"{recycled_powder_g:,.1f}"),
                            ("تاریخ بازیافت (شمسی)", rec_date),
                            ("توضیحات و ملاحظات غربال‌گری", rec_notes or "ندارد")
                        ]
                    }
                    excel_rec_bytes = build_form_layout_excel(export_rec_sec, title=f"فرم بازیافت پودر - {recycled_batch_code or 'جدید'}")
                    st.download_button(
                        label="📥 دانلود خروجی اکسل",
                        data=excel_rec_bytes,
                        file_name=f"recycled_powder_{recycled_batch_code or 'record'}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="btn_down_rec_form_styled"
                    )

        with sub_tab3:
            st.subheader("🏭 پودرهای خریداری شده از نورا (ثبت سریع و ویرایش)")
            mode_nora = st.radio("", ["ثبت پودر جدید نورا", "ویرایش پودر موجود نورا"], horizontal=True, key="mode_nora")
            edit_nora_data = {}
            if mode_nora == "ویرایش پودر موجود نورا":
                if not nora_df.empty:
                    nora_codes = nora_df['powder_code'].tolist()
                    sel_nora_code = st.selectbox("انتخاب کد پودر نورا جهت ویرایش:", nora_codes, key="sel_nora_code")
                    edit_nora_data = nora_df[nora_df['powder_code'] == sel_nora_code].iloc[0].to_dict()
                else:
                    st.warning("هیچ پودری از نورا ثبت نشده است.")

            nc1, nc2 = st.columns(2)
            with nc1:
                def_nora_code = edit_nora_data.get('powder_code', '') if mode_nora == "ویرایش پودر موجود نورا" else ""
                nora_powder_code = st.text_input("کد/شماره ظرف پودر نورا (مانند: NORA-PWD-01)", value=def_nora_code, disabled=(mode_nora == "ویرایش پودر موجود نورا"))
                
                n_mat_opts = ["Steel 316", "Ti6Al4V", "Inconel 718", "Hastelloy X", "سایر"]
                def_nmat_idx = n_mat_opts.index(edit_nora_data['material']) if mode_nora == "ویرایش پودر موجود نورا" and edit_nora_data.get('material') in n_mat_opts else 0
                nora_material = st.selectbox("نوع متریال / جنس پودر", n_mat_opts, index=def_nmat_idx, key="nora_mat")
            with nc2:
                def_nwt = float(edit_nora_data.get('weight_g', 10000.0))
                nora_weight_g = st.number_input("مقدار / وزن پودر (گرم)", min_value=0.0, value=def_nwt, key="nora_wt")
                
                def_ndt = format_jalali_date(edit_nora_data.get('date', '')) if mode_nora == "ویرایش پودر موجود نورا" else get_today_jalali()
                nora_date = st.text_input("تاریخ ورود/تحویل (هجری شمسی)", value=def_ndt, key="nora_dt")
                
            n_btn1, n_btn2 = st.columns([3, 2])
            with n_btn1:
                nora_submit = st.button("💾 ثبت پودر نورا در بایگانی")
                if nora_submit:
                    target_ncode = def_nora_code if mode_nora == "ویرایش پودر موجود نورا" else nora_powder_code
                    if target_ncode:
                        c = conn.cursor()
                        c.execute("""INSERT OR REPLACE INTO nora_powders (powder_code, material, weight_g, date) VALUES (?, ?, ?, ?)""",
                                  (target_ncode, nora_material, nora_weight_g, nora_date))
                        conn.commit()
                        st.success(f"پودر نورا با کد {target_ncode} با موفقیت ذخیره شد.")
                        st.rerun()
                    else:
                        st.error("لطفاً کد/شماره ظرف پودر را وارد کنید.")
            with n_btn2:
                export_nora_sec = {
                    "مشخصات پودر خریداری شده از نورا": [
                        ("کد / شماره ظرف پودر نورا", def_nora_code or nora_powder_code or "---"),
                        ("جنس / متریال پودر", nora_material),
                        ("وزن / مقدار پودر (گرم)", f"{nora_weight_g:,.1f}"),
                        ("تاریخ ورود / تحویل (شمسی)", nora_date)
                    ]
                }
                excel_nora_bytes = build_form_layout_excel(export_nora_sec, title=f"فرم پودر نورا - {def_nora_code or nora_powder_code or 'جدید'}")
                st.download_button(
                    label="📥 دانلود خروجی اکسل",
                    data=excel_nora_bytes,
                    file_name=f"nora_powder_{def_nora_code or nora_powder_code or 'record'}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_down_nora_form_styled"
                )

        conn.close()

    # ---------------------------------------------------------
    # ۲. فرم تولید
    # ---------------------------------------------------------
    elif choice == "🏭 فرم تولید":
        st.header("🏭 فرم رکورد تولید")
        
        conn = get_db_connection()
        powders_list = pd.read_sql_query("SELECT powder_code FROM powders", conn)['powder_code'].tolist()
        nora_powders_list = pd.read_sql_query("SELECT powder_code FROM nora_powders", conn)['powder_code'].tolist()
        all_powders = list(set(powders_list + nora_powders_list))
        
        st.subheader("حالت")
        mode = st.radio("", ["ثبت قطعه جدید", "ویرایش قطعه موجود"], horizontal=True, key="prod_mode")
        
        edit_data = {}
        if mode == "ویرایش قطعه موجود":
            existing_parts = pd.read_sql_query("SELECT part_code FROM production", conn)['part_code'].tolist()
            if existing_parts:
                selected_edit_code = st.selectbox("انتخاب قطعه جهت ویرایش:", existing_parts)
                edit_data = pd.read_sql_query("SELECT * FROM production WHERE part_code=?", conn, params=(selected_edit_code,)).iloc[0].to_dict()
            else:
                st.warning("هنوز هیچ قطعه‌ای جهت ویرایش ثبت نشده است.")

        if not all_powders:
            st.warning("جهت ثبت فرم تولید، ابتدا باید حداقل یک پودر در بخش ۱ ثبت شده باشد.")
        else:
            col1, col2, col3 = st.columns(3)
            with col1:
                default_part_code = edit_data.get("part_code", "") if mode == "ویرایش قطعه موجود" else ""
                part_code = st.text_input("کد قطعه (شناسه یکتا)", value=default_part_code, disabled=(mode == "ویرایش قطعه موجود"))
                part_name = st.text_input("نام قطعه", value=edit_data.get("part_name", ""))
            with col2:
                default_powder_index = all_powders.index(edit_data["powder_code"]) if mode == "ویرایش قطعه موجود" and edit_data.get("powder_code") in all_powders else 0
                powder_code = st.selectbox("شماره ظرف پودر مصرفی", all_powders, index=default_powder_index)
                
                machine_options = ["M120", "M300", "سایر"]
                default_machine_index = machine_options.index(edit_data["machine_model"]) if mode == "ویرایش قطعه موجود" and edit_data.get("machine_model") in machine_options else 0
                machine_model = st.selectbox("مدل دستگاه", machine_options, index=default_machine_index)
            with col3:
                quantity = st.number_input("تعداد روی صفحه ساخت", min_value=1, value=int(edit_data.get("quantity", 1)))
                def_prod_dt = format_jalali_date(edit_data.get("date", '')) if mode == "ویرایش قطعه موجود" else get_today_jalali()
                date_str = st.text_input("تاریخ ساخت (هجری شمسی)", value=def_prod_dt)
                
            st.markdown("---")
            st.subheader("⏱️ ۱- زمان آماده‌سازی و ساخت (فرمت ساعت:دقیقه مانند 08:30)")
            tc1, tc2, tc3 = st.columns(3)
            with tc1:
                def_build_time_str = hours_to_time_str(edit_data.get("build_time_hrs", 10.0)) if mode == "ویرایش قطعه موجود" else "10:00"
                build_time_input = st.text_input("زمان تولید (ساعت:دقیقه)", value=def_build_time_str)
                
                def_downtime_str = hours_to_time_str(edit_data.get("downtime_hrs", 0.0)) if mode == "ویرایش قطعه موجود" else "00:00"
                downtime_input = st.text_input("زمان توقف حین ساخت (ساعت:دقیقه)", value=def_downtime_str)
            with tc2:
                start_date = st.text_input("تاریخ شروع (هجری شمسی)", value=format_jalali_date(edit_data.get("start_date", date_str)))
                start_time = st.text_input("ساعت شروع (مانند 08:00)", value=edit_data.get("start_time", "08:00"))
                end_date = st.text_input("تاریخ پایان (هجری شمسی)", value=format_jalali_date(edit_data.get("end_date", date_str)))
                end_time = st.text_input("ساعت پایان (مانند 16:00)", value=edit_data.get("end_time", "16:00"))
            with tc3:
                def_setup_str = hours_to_time_str(edit_data.get("setup_time_hrs", 1.0)) if mode == "ویرایش قطعه موجود" else "01:00"
                setup_time_input = st.text_input("زمان آماده‌سازی دستگاه (ساعت:دقیقه)", value=def_setup_str)
                
                def_clean_str = hours_to_time_str(edit_data.get("cleaning_time_hrs", 0.75)) if mode == "ویرایش قطعه موجود" else "00:45"
                cleaning_time_input = st.text_input("زمان تمیزکاری دستگاه (ساعت:دقیقه)", value=def_clean_str)
                
            st.markdown("---")
            st.subheader("⚖️ ۲- پارامترهای متریال و وزن")
            mc1, mc2 = st.columns(2)
            with mc1:
                input_powder_g = st.number_input("پودر ورودی به دستگاه (گرم)", min_value=0.0, value=float(edit_data.get("input_powder_g", 0.0)))
                waste_powder_g = st.number_input("پودر مصرف شده غیر قابل بازیافت (گرم)", min_value=0.0, value=float(edit_data.get("waste_powder_g", 0.0)))
                part_with_support_g = st.number_input("وزن قطعه با ساپورت (گرم)", min_value=0.0, value=float(edit_data.get("part_with_support_g", 0.0)))
                final_part_g = st.number_input("وزن قطعه نهایی (گرم)", min_value=0.0, value=float(edit_data.get("final_part_g", 0.0)))
                filter_pct = st.number_input("درصد فیلتر دستگاه (%)", min_value=0.0, max_value=100.0, value=float(edit_data.get("filter_percentage", 0.0)))
            with mc2:
                plate_code = st.text_input("کد صفحه ساخت", value=edit_data.get("build_plate_code", ""))
                plate_init_wt = st.number_input("وزن اولیه صفحه ساخت (گرم)", min_value=0.0, value=float(edit_data.get("build_plate_init_wt_g", 0.0)))
                plate_post_wt = st.number_input("وزن صفحه ساخت پس از پرداخت (گرم)", min_value=0.0, value=float(edit_data.get("build_plate_post_wt_g", 0.0)))
                
            p_btn1, p_btn2 = st.columns([3, 2])
            with p_btn1:
                submitted = st.button("💾 ذخیره تغییرات / ثبت رکورد تولید")
                if submitted:
                    target_code = default_part_code if mode == "ویرایش قطعه موجود" else part_code
                    if target_code:
                        build_time_hrs_dec = time_to_hours(build_time_input)
                        downtime_hrs_dec = time_to_hours(downtime_input)
                        setup_time_hrs_dec = time_to_hours(setup_time_input)
                        cleaning_time_hrs_dec = time_to_hours(cleaning_time_input)
                        
                        c = conn.cursor()
                        c.execute("""INSERT OR REPLACE INTO production 
                            (part_code, part_name, powder_code, quantity, machine_model, date,
                             build_time_hrs, downtime_hrs, start_date, start_time, end_date, end_time,
                             setup_time_hrs, cleaning_time_hrs, input_powder_g, waste_powder_g,
                             part_with_support_g, final_part_g, filter_percentage, build_plate_code,
                             build_plate_init_wt_g, build_plate_post_wt_g)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (target_code, part_name, powder_code, quantity, machine_model, date_str,
                             build_time_hrs_dec, downtime_hrs_dec, start_date, start_time, end_date, end_time,
                             setup_time_hrs_dec, cleaning_time_hrs_dec, input_powder_g, waste_powder_g,
                             part_with_support_g, final_part_g, filter_pct, plate_code,
                             plate_init_wt, plate_post_wt))
                        conn.commit()
                        st.success(f"اطلاعات تولید قطعه {target_code} با موفقیت ذخیره شد.")
                        st.rerun()
                    else:
                        st.error("لطفاً کد قطعه را مشخص کنید.")
            with p_btn2:
                target_pcode_exp = default_part_code if mode == "ویرایش قطعه موجود" else part_code
                prod_export_sections = {
                    "مشخصات عمومی قطعه و دستگاه": [
                        ("کد قطعه (شناسه یکتا)", target_pcode_exp or "---"),
                        ("نام قطعه", part_name or "---"),
                        ("شماره ظرف پودر مصرفی", powder_code),
                        ("مدل دستگاه", machine_model),
                        ("تعداد روی صفحه ساخت", quantity),
                        ("تاریخ ساخت (شمسی)", date_str)
                    ],
                    "۱- پارامترهای زمان‌بندی ساخت و آماده‌سازی": [
                        ("زمان تولید (ساعت:دقیقه)", build_time_input),
                        ("زمان توقف حین ساخت", downtime_input),
                        ("تاریخ و ساعت شروع", f"{start_date} - {start_time}"),
                        ("تاریخ و ساعت پایان", f"{end_date} - {end_time}"),
                        ("زمان آماده‌سازی دستگاه", setup_time_input),
                        ("زمان تمیزکاری دستگاه", cleaning_time_input)
                    ],
                    "۲- پارامترهای متریال، وزن و صفحه ساخت": [
                        ("پودر ورودی به دستگاه (گرم)", f"{input_powder_g:,.1f}"),
                        ("پودر غیرقابل بازیافت (گرم)", f"{waste_powder_g:,.1f}"),
                        ("وزن قطعه با ساپورت (گرم)", f"{part_with_support_g:,.1f}"),
                        ("وزن قطعه نهایی (گرم)", f"{final_part_g:,.1f}"),
                        ("درصد فیلتر دستگاه (%)", f"{filter_pct}%"),
                        ("کد صفحه ساخت", plate_code or "---"),
                        ("وزن اولیه صفحه ساخت (گرم)", f"{plate_init_wt:,.1f}"),
                        ("وزن صفحه پس از پرداخت (گرم)", f"{plate_post_wt:,.1f}")
                    ]
                }
                excel_prod_form_bytes = build_form_layout_excel(prod_export_sections, title=f"کاربرگ رکورد تولید - قطعه {target_pcode_exp or 'جدید'}")
                st.download_button(
                    label="📥 دانلود خروجی اکسل",
                    data=excel_prod_form_bytes,
                    file_name=f"production_form_{target_pcode_exp or 'record'}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_down_prod_form_styled"
                )

        conn.close()

    # ---------------------------------------------------------
    # ۳. کنترل کیفیت
    # ---------------------------------------------------------
    elif choice == "❇️ کنترل کیفیت":
        st.header("🔬 فرم کنترل کیفیت")
        
        conn = get_db_connection()
        parts_list = pd.read_sql_query("SELECT part_code, part_name FROM production", conn)
        
        if parts_list.empty:
            st.warning("هیچ قطعه‌ای در بخش فرم تولید ثبت نشده است.")
        else:
            selected_part = st.selectbox("انتخاب کد قطعه جهت ارزیابی / ویرایش کیفیت:", parts_list['part_code'])
            part_info = pd.read_sql_query("SELECT * FROM production WHERE part_code=?", conn, params=(selected_part,)).iloc[0]
            
            existing_qc = pd.read_sql_query("SELECT * FROM qc WHERE part_code=?", conn, params=(selected_part,))
            saved_qc_data = {}
            saved_inspector, saved_engineer, saved_manager = "", "", ""
            saved_qc_date = get_today_jalali()
            
            if not existing_qc.empty:
                st.info("⚠️ برای این قطعه قبلاً فرم QC ثبت شده است؛ مقادیر و نوت‌های قبلی بارگذاری شدند و می‌توانید آن‌ها را ویرایش کنید.")
                qc_row = existing_qc.iloc[0]
                saved_inspector = qc_row.get('qc_inspector', '')
                saved_engineer = qc_row.get('qc_engineer', '')
                saved_manager = qc_row.get('qa_manager', '')
                saved_qc_date = format_jalali_date(qc_row.get('date', ''))
                raw_qc_json = qc_row.get('qc_checks_json', '')
                if raw_qc_json:
                    try:
                        saved_qc_data = json.loads(raw_qc_json)
                    except Exception:
                        pass
            
            st.info(f"**نام قطعه:** {part_info['part_name']} | **دستگاه:** {part_info['machine_model']} | **شماره ظرف پودر:** {part_info['powder_code']}")
            
            qc_date_input = st.text_input("تاریخ ثبت بازرسی (هجری شمسی)", value=saved_qc_date)
            
            st.subheader("📋 تست‌ها و بازرسی‌های کیفی (شامل نتیجه و یادداشت‌های تفکیکی)")
            tests = [
                ("1.1.1 ظاهر سطح و عیوب قابل رؤیت", "چشمی"),
                ("1.1.2 کیفیت حذف ساپورت", "چشمی"),
                ("1.1.3 زبری سطح (Ra)", "ابزار زبری‌سنج"),
                ("2.1.1 ابعاد بحرانی", "کولیس/CMM"),
                ("2.1.2 تختی، هم‌محوری، عمودیت و GD&T", "CMM/ژئومتریک"),
                ("2.1.3 انطباق با مدل CAD", "اسکن سه بعدی"),
                ("3.1.1 تخلخل/ترک/ناپیوستگی", "NDT"),
                ("4.1.1 استحکام کششی / تسلیم", "تست مکانیکی"),
                ("4.1.2 سختی", "سختی‌سنج"),
                ("4.1.3 چگالی نسبی / تخلخل", "دانسیته‌سنج"),
                ("4.1.4 ریزساختار و کیفیت ذوب", "متالوگرافی")
            ]
            
            qc_data = {}
            for t_title, t_type in tests:
                q1, q2, q3 = st.columns([3, 2, 4])
                item_prev = saved_qc_data.get(t_title, {})
                prev_res = item_prev.get('result', 'تایید') if isinstance(item_prev, dict) else 'تایید'
                prev_note = item_prev.get('note', '') if isinstance(item_prev, dict) else ''
                
                with q1: st.write(f"**{t_title}**")
                with q2: res = st.selectbox("نتیجه", ["تایید", "رد"], index=0 if prev_res=="تایید" else 1, key=f"qc_sel_{t_title}")
                with q3: note = st.text_input("ملاحظات", value=prev_note, key=f"qc_n_{t_title}")
                qc_data[t_title] = {"result": res, "type": t_type, "note": note}
                
            st.markdown("---")
            st.subheader("👥 مسئولین و تاییدکنندگان")
            sc1, sc2, sc3 = st.columns(3)
            with sc1: inspector = st.text_input("بازرس کنترل کیفیت", value=saved_inspector)
            with sc2: engineer = st.text_input("مسئول فنی / مهندسی کیفیت", value=saved_engineer)
            with sc3: manager = st.text_input("مدیر تضمین کیفیت", value=saved_manager)
            
            qc_btn1, qc_btn2 = st.columns([3, 2])
            with qc_btn1:
                qc_submit = st.button("💾 ثبت / به‌روزرسانی نهایی فرم QC")
                if qc_submit:
                    c = conn.cursor()
                    c.execute("""INSERT OR REPLACE INTO qc 
                                 (part_code, part_name, material, machine_model, date, qc_checks_json, qc_inspector, qc_engineer, qa_manager)
                                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                              (selected_part, part_info['part_name'], part_info['powder_code'], part_info['machine_model'],
                               qc_date_input, json.dumps(qc_data, ensure_ascii=False), inspector, engineer, manager))
                    conn.commit()
                    st.success("نتایج ارزیابی QC با موفقیت ذخیره شد.")
                    st.rerun()
            with qc_btn2:
                export_qc_sec = {
                    "مشخصات قطعه مورد ارزیابی": [
                        ("کد قطعه", selected_part),
                        ("نام قطعه", part_info['part_name']),
                        ("مدل دستگاه ساخت", part_info['machine_model']),
                        ("شماره ظرف پودر مصرفی", part_info['powder_code']),
                        ("تاریخ بازرسی (شمسی)", qc_date_input),
                        ("بازرس کنترل کیفیت", inspector or "---"),
                        ("مسئول فنی / مهندسی کیفیت", engineer or "---"),
                        ("مدیر تضمین کیفیت", manager or "---")
                    ],
                    "چک‌لیست آزمون‌ها، نتایج بازرسی و ملاحظات فنی": qc_data
                }
                excel_qc_form_bytes = build_form_layout_excel(export_qc_sec, title=f"گزارش کنترل کیفیت - قطعه {selected_part}")
                st.download_button(
                    label="📥 دانلود خروجی اکسل",
                    data=excel_qc_form_bytes,
                    file_name=f"qc_report_{selected_part}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_down_qc_form_styled"
                )

        conn.close()

    # ---------------------------------------------------------
    # ۴. محاسبه‌گر هزینه
    # ---------------------------------------------------------
    elif choice == "💰 محاسبه‌گر هزینه":
        st.header("💰 محاسبه‌گر بهای تمام شده و قیمت فروش")
        
        SYSTEM_RATES = load_system_rates()
        
        RATES = {
            "powder_price_per_kg": {
                "Steel 316": SYSTEM_RATES["powder_price_Steel_316"],
                "Ti6Al4V": SYSTEM_RATES["powder_price_Ti6Al4V"],
                "Inconel 718": SYSTEM_RATES["powder_price_Inconel_718"],
                "Hastelloy X": SYSTEM_RATES["powder_price_Hastelloy_X"]
            },
            "machine_depreciation_hr": {
                "M120": SYSTEM_RATES["machine_depr_M120"],
                "M300": SYSTEM_RATES["machine_depr_M300"]
            },
            "power_kw_hr": {
                "M120": SYSTEM_RATES["power_kw_M120"],
                "M300": SYSTEM_RATES["power_kw_M300"]
            },
            "wages": {
                "designer_hr": SYSTEM_RATES["wage_designer"],
                "operator_hr": SYSTEM_RATES["wage_operator"],
                "qc_hr": SYSTEM_RATES["wage_qc"]
            },
            "argon_hr": SYSTEM_RATES["argon_rate"],
            "electricity_kwh": SYSTEM_RATES["electricity_rate"],
            "post_process_hr": SYSTEM_RATES["post_process_rate"],
            "qc_fixed_cost": SYSTEM_RATES["qc_fixed_cost"],
            "ventilation_hr": SYSTEM_RATES["ventilation_rate"],
            "chiller_hr": SYSTEM_RATES["chiller_rate"],
            "density": {"Steel 316": 8.0, "Ti6Al4V": 4.43, "Inconel 718": 8.19, "Hastelloy X": 8.22}
        }
        
        conn = get_db_connection()
        parts_list = pd.read_sql_query("SELECT part_code, part_name FROM production", conn)
        
        selected_part = st.selectbox("فراخوانی قطعه از بخش تولید (یا وارد کردن قطعه جدید)", ["جدید"] + list(parts_list['part_code']))
        
        def_part_name = ""
        def_machine = "M300"
        def_print_time_str = "10:00"
        def_vol = 50.0
        def_sup_vol = 10.0
        def_parts_plate = 1
        def_design_time_str = "02:00"
        def_post_time_str = "03:00"
        def_overhead = 35.0
        def_pwd_type = "Steel 316"
        
        if selected_part != "جدید":
            prev_cost_row = pd.read_sql_query("SELECT * FROM cost_calculator WHERE part_code=?", conn, params=(selected_part,))
            if not prev_cost_row.empty:
                c_row = prev_cost_row.iloc[0]
                def_part_name = c_row['part_name']
                def_machine = c_row['machine_type']
                def_print_time_str = hours_to_time_str(c_row['print_time_hrs'])
                def_vol = float(c_row['volume_cm3'])
                def_sup_vol = float(c_row['support_volume_cm3'])
                def_parts_plate = int(c_row['parts_on_plate'])
                def_design_time_str = hours_to_time_str(c_row['design_time_hrs'])
                def_post_time_str = hours_to_time_str(c_row['post_process_time_hrs'])
                def_overhead = float(c_row['overhead_pct'])
                def_pwd_type = c_row['powder_type']
                st.info("⚠️ محاسبات مالی قبلی این قطعه لود شد و می‌توانید آن‌ها را تغییر دهید.")
            else:
                p_row = pd.read_sql_query("SELECT * FROM production WHERE part_code=?", conn, params=(selected_part,))
                if not p_row.empty:
                    p_info = p_row.iloc[0]
                    def_part_name = p_info['part_name']
                    def_machine = p_info['machine_model'] if p_info['machine_model'] in ["M120", "M300"] else "M300"
                    def_print_time_str = hours_to_time_str(p_info['build_time_hrs'])

        st.subheader("📥 مشخصات فنی و ورودی‌های قطعه")
        c1, c2, c3 = st.columns(3)
        with c1:
            p_code = st.text_input("شناسه/کد قطعه", value=selected_part if selected_part != "جدید" else "")
            p_name = st.text_input("نام قطعه", value=def_part_name)
            
            pwd_type_list = list(RATES["density"].keys())
            def_ptype_idx = pwd_type_list.index(def_pwd_type) if def_pwd_type in pwd_type_list else 0
            powder_type = st.selectbox("نوع پودر فلزی", pwd_type_list, index=def_ptype_idx)
        with c2:
            vol_cm3 = st.number_input("حجم قطعه (cm3)", min_value=0.0, value=def_vol)
            sup_vol_cm3 = st.number_input("حجم ساپورت (cm3)", min_value=0.0, value=def_sup_vol)
            machine_type = st.selectbox("نوع دستگاه", ["M120", "M300"], index=1 if def_machine=="M300" else 0)
        with c3:
            parts_on_plate = st.number_input("تعداد قطعات روی صفحه", min_value=1, value=def_parts_plate)
            print_time_input = st.text_input("زمان کل چاپ (ساعت:دقیقه)", value=def_print_time_str)
            design_time_input = st.text_input("زمان طراحی (ساعت:دقیقه)", value=def_design_time_str)
            post_time_input = st.text_input("زمان پرداخت‌کاری (ساعت:دقیقه)", value=def_post_time_str)
            overhead_pct = st.number_input("ضریب سربار (%)", min_value=0.0, value=def_overhead)

        print_time_hrs = time_to_hours(print_time_input)
        design_time_hrs = time_to_hours(design_time_input)
        post_time_hrs = time_to_hours(post_time_input)

        density = RATES["density"][powder_type]
        net_weight_g = vol_cm3 * density
        support_weight_g = sup_vol_cm3 * density
        total_weight_kg = (net_weight_g + support_weight_g) / 1000.0
        
        cost_powder = total_weight_kg * RATES["powder_price_per_kg"][powder_type]
        cost_argon = print_time_hrs * RATES["argon_hr"]
        cost_depreciation = print_time_hrs * RATES["machine_depreciation_hr"][machine_type]
        cost_power = print_time_hrs * RATES["power_kw_hr"][machine_type] * RATES["electricity_kwh"]
        cost_engineering = design_time_hrs * RATES["wages"]["designer_hr"]
        cost_operator = print_time_hrs * RATES["wages"]["operator_hr"]
        cost_post_process = post_time_hrs * RATES["post_process_hr"]
        cost_qc = RATES["qc_fixed_cost"]
        cost_ventilation = print_time_hrs * RATES["ventilation_hr"]
        cost_chiller = print_time_hrs * RATES["chiller_hr"]
        
        total_production_cost = (cost_powder + cost_argon + cost_depreciation + cost_power + 
                                 cost_engineering + cost_operator + cost_post_process + 
                                 cost_qc + cost_ventilation + cost_chiller)
        
        overhead_cost = total_production_cost * (overhead_pct / 100.0)
        final_price = total_production_cost + overhead_cost
        
        st.markdown("---")
        st.subheader("📊 تفکیک و خلاصه هزینه‌های برآورد شده")
        
        mc1, mc2, mc3 = st.columns(3)
        mc1.metric("وزن خالص قطعه", f"{net_weight_g:.1f} گرم")
        mc2.metric("بهای تمام شده کل", f"{total_production_cost:,.0f} ریال")
        mc3.metric("قیمت نهایی قابل ارائه به مشتری", f"{final_price:,.0f} ریال")
        
        cost_btn1, cost_btn2 = st.columns([3, 2])
        with cost_btn1:
            save_cost_calc = st.button("💾 ذخیره / به‌روزرسانی برآورد قیمت")
            if save_cost_calc:
                if p_code:
                    c = conn.cursor()
                    c.execute("""INSERT OR REPLACE INTO cost_calculator 
                        (part_code, part_name, powder_type, volume_cm3, net_weight_g, support_volume_cm3, support_weight_g,
                         machine_type, parts_on_plate, print_time_hrs, design_time_hrs, post_process_time_hrs, overhead_pct,
                         powder_cost_total, argon_cost_total, depreciation_cost_total, power_cost_total, engineering_cost_total,
                         operator_cost_total, post_process_cost_total, qc_cost_total, utility_ventilation, utility_chiller,
                         total_production_cost, overhead_cost, final_price)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (p_code, p_name, powder_type, vol_cm3, net_weight_g, sup_vol_cm3, support_weight_g,
                         machine_type, parts_on_plate, print_time_hrs, design_time_hrs, post_time_hrs, overhead_pct,
                         cost_powder, cost_argon, cost_depreciation, cost_power, cost_engineering,
                         cost_operator, cost_post_process, cost_qc, cost_ventilation, cost_chiller,
                         total_production_cost, overhead_cost, final_price))
                    conn.commit()
                    st.success("محاسبه هزینه با موفقیت ذخیره شد.")
                    st.rerun()
                else:
                    st.error("لطفاً شناسه قطعه را مشخص کنید.")
        with cost_btn2:
            export_cost_sec = {
                "ورودی‌های مهندسی و مشخصات فنی قطعه": [
                    ("کد قطعه", p_code or "---"),
                    ("نام قطعه", p_name or "---"),
                    ("نوع پودر فلزی", powder_type),
                    ("مدل دستگاه انتخابی", machine_type),
                    ("حجم قطعه (cm3)", f"{vol_cm3:.2f}"),
                    ("حجم ساپورت (cm3)", f"{sup_vol_cm3:.2f}"),
                    ("وزن خالص قطعه (گرم)", f"{net_weight_g:.1f}"),
                    ("وزن ساپورت (گرم)", f"{support_weight_g:.1f}"),
                    ("تعداد روی صفحه", parts_on_plate),
                    ("زمان کل چاپ (ساعت:دقیقه)", print_time_input),
                    ("زمان طراحی (ساعت:دقیقه)", design_time_input),
                    ("زمان پرداخت (ساعت:دقیقه)", post_time_input)
                ],
                "تفکیک هزینه‌های مستقیم، سربار و قیمت فروش": [
                    ("هزینه پودر مصرفی (ریال)", f"{cost_powder:,.0f}"),
                    ("هزینه گاز آرگون (ریال)", f"{cost_argon:,.0f}"),
                    ("استهلاک دستگاه ساخت (ریال)", f"{cost_depreciation:,.0f}"),
                    ("هزینه برق مصرفی (ریال)", f"{cost_power:,.0f}"),
                    ("دستمزد مهندسی و طراحی (ریال)", f"{cost_engineering:,.0f}"),
                    ("دستمزد اپراتور دستگاه (ریال)", f"{cost_operator:,.0f}"),
                    ("هزینه پرداخت‌کاری سطحی (ریال)", f"{cost_post_process:,.0f}"),
                    ("هزینه تست‌های کیفی QC (ریال)", f"{cost_qc:,.0f}"),
                    ("هزینه تهویه و چیلر (ریال)", f"{cost_ventilation + cost_chiller:,.0f}"),
                    ("ضریب سربار کارگاهی (%)", f"{overhead_pct}%"),
                    ("بهای تمام شده کل (ریال)", f"{total_production_cost:,.0f}"),
                    ("قیمت نهایی قابل ارائه به مشتری (ریال)", f"{final_price:,.0f}")
                ]
            }
            excel_cost_form_bytes = build_form_layout_excel(export_cost_sec, title=f"برآورد مالی و بهای تمام شده - قطعه {p_code or 'جدید'}")
            st.download_button(
                label="📥 دانلود خروجی اکسل",
                data=excel_cost_form_bytes,
                file_name=f"cost_estimate_{p_code or 'record'}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_down_cost_form_styled"
            )

        conn.close()

    # ---------------------------------------------------------
    # ۵. تنظیمات نرخ‌ها
    # ---------------------------------------------------------
    elif choice == "⚙️ تنظیمات نرخ‌ها":
        st.header("⚙️ تنظیمات و ویرایش نرخ‌های پایه محاسبات")
        st.info("تغییرات در این بخش بلافاصله در «محاسبه‌گر هزینه» و فرآیندهای مالی اعمال می‌شوند.")
        
        current_rates = load_system_rates()
        
        st.subheader("۱. قیمت پودرهای فلزی (ریال به ازای هر کیلوگرم)")
        c1, c2 = st.columns(2)
        with c1:
            r_steel = st.number_input("پودر Steel 316", min_value=0.0, value=float(current_rates.get("powder_price_Steel_316", 120000000)))
            r_ti = st.number_input("پودر Ti6Al4V", min_value=0.0, value=float(current_rates.get("powder_price_Ti6Al4V", 500000000)))
        with c2:
            r_inconel = st.number_input("پودر Inconel 718", min_value=0.0, value=float(current_rates.get("powder_price_Inconel_718", 300000000)))
            r_hastelloy = st.number_input("پودر Hastelloy X", min_value=0.0, value=float(current_rates.get("powder_price_Hastelloy_X", 400000000)))

        st.markdown("---")
        st.subheader("۲. نرخ ساعتی استهلاک دستگاه‌ها (ریال / ساعت)")
        m1, m2 = st.columns(2)
        with m1:
            r_depr_m120 = st.number_input("نرخ استهلاک دستگاه M120", min_value=0.0, value=float(current_rates.get("machine_depr_M120", 1250000)))
        with m2:
            r_depr_m300 = st.number_input("نرخ استهلاک دستگاه M300", min_value=0.0, value=float(current_rates.get("machine_depr_M300", 3125000)))

        st.markdown("---")
        st.subheader("۳. دستمزدهای نیروی انسانی (ریال / نفرساعت)")
        w1, w2, w3 = st.columns(3)
        with w1:
            r_wage_designer = st.number_input("دستمزد طراح مهندسی", min_value=0.0, value=float(current_rates.get("wage_designer", 2500000)))
        with w2:
            r_wage_operator = st.number_input("دستمزد اپراتور دستگاه", min_value=0.0, value=float(current_rates.get("wage_operator", 1870000)))
        with w3:
            r_wage_qc = st.number_input("دستمزد مسئول کنترل کیفیت", min_value=0.0, value=float(current_rates.get("wage_qc", 2180000)))

        st.markdown("---")
        st.subheader("۴. هزینه‌های جانبی، انرژی و مصارف کارگاهی")
        u1, u2, u3 = st.columns(3)
        with u1:
            r_argon = st.number_input("نرخ گاز آرگون (ریال/ساعت)", min_value=0.0, value=float(current_rates.get("argon_rate", 300000)))
            r_electricity = st.number_input("نرخ برق مصرفی (ریال/کیلووات‌ساعت)", min_value=0.0, value=float(current_rates.get("electricity_rate", 50000)))
        with u2:
            r_post_process = st.number_input("عملیات پرداخت سطحی (ریال/ماشین‌ساعت)", min_value=0.0, value=float(current_rates.get("post_process_rate", 500000)))
            r_qc_fixed = st.number_input("تست‌های QC ثابت (ریال/قطعه)", min_value=0.0, value=float(current_rates.get("qc_fixed_cost", 40000000)))
        with u3:
            r_ventilation = st.number_input("سیستم تهویه (ریال/ساعت)", min_value=0.0, value=float(current_rates.get("ventilation_rate", 200000)))
            r_chiller = st.number_input("آب خنک‌کاری چیلر (ریال/ساعت)", min_value=0.0, value=float(current_rates.get("chiller_rate", 2500)))

        rate_btn1, rate_btn2 = st.columns([3, 2])
        with rate_btn1:
            save_rates_submit = st.button("💾 ذخیره تغییرات نرخ‌ها در پایگاه داده")
            if save_rates_submit:
                conn = get_db_connection()
                c = conn.cursor()
                updated_pairs = [
                    ("powder_price_Steel_316", r_steel),
                    ("powder_price_Ti6Al4V", r_ti),
                    ("powder_price_Inconel_718", r_inconel),
                    ("powder_price_Hastelloy_X", r_hastelloy),
                    ("machine_depr_M120", r_depr_m120),
                    ("machine_depr_M300", r_depr_m300),
                    ("wage_designer", r_wage_designer),
                    ("wage_operator", r_wage_operator),
                    ("wage_qc", r_wage_qc),
                    ("argon_rate", r_argon),
                    ("electricity_rate", r_electricity),
                    ("post_process_rate", r_post_process),
                    ("qc_fixed_cost", r_qc_fixed),
                    ("ventilation_rate", r_ventilation),
                    ("chiller_rate", r_chiller)
                ]
                for k, val in updated_pairs:
                    c.execute("INSERT OR REPLACE INTO system_settings (key, value) VALUES (?, ?)", (k, val))
                conn.commit()
                conn.close()
                st.success("تمام نرخ‌های جدید با موفقیت ذخیره شدند و در بخش «محاسبه‌گر هزینه» اعمال گردیدند.")
                st.rerun()
        with rate_btn2:
            export_rate_sec = {
                "نرخ‌های پایه متریال و پودر (ریال/کیلوگرم)": [
                    ("پودر Steel 316", f"{r_steel:,.0f}"),
                    ("پودر Ti6Al4V", f"{r_ti:,.0f}"),
                    ("پودر Inconel 718", f"{r_inconel:,.0f}"),
                    ("پودر Hastelloy X", f"{r_hastelloy:,.0f}")
                ],
                "نرخ استهلاک دستگاه‌ها و دستمزدها (ریال/ساعت)": [
                    ("دستگاه M120", f"{r_depr_m120:,.0f}"),
                    ("دستگاه M300", f"{r_depr_m300:,.0f}"),
                    ("دستمزد طراح", f"{r_wage_designer:,.0f}"),
                    ("دستمزد اپراتور", f"{r_wage_operator:,.0f}"),
                    ("دستمزد QC", f"{r_wage_qc:,.0f}")
                ],
                "مصارف، انرژی و خدمات جانبی": [
                    ("گاز آرگون (ریال/ساعت)", f"{r_argon:,.0f}"),
                    ("برق (ریال/کیلووات)", f"{r_electricity:,.0f}"),
                    ("پرداخت‌کاری (ریال/ساعت)", f"{r_post_process:,.0f}"),
                    ("تست ثابت QC (ریال/قطعه)", f"{r_qc_fixed:,.0f}"),
                    ("تهویه و چیلر (ریال/ساعت)", f"{r_ventilation + r_chiller:,.0f}")
                ]
            }
            excel_rates_form_bytes = build_form_layout_excel(export_rate_sec, title="جدول تنظیمات نرخ‌های پایه سامانه SLM")
            st.download_button(
                label="📥 دانلود خروجی اکسل",
                data=excel_rates_form_bytes,
                file_name="system_rates_form.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_down_rates_top"
            )

    # ---------------------------------------------------------
    # ۶. بایگانی و گزارش‌گیری اکسل (پرونده جامع فرم‌محور)
    # ---------------------------------------------------------
    elif choice == "🔍 بایگانی":
        st.header("🔍 بایگانی جامع، استعلام پرونده‌ها و مدیریت رکوردها")
        
        conn = get_db_connection()
        
        st.subheader("🔎 استعلام شناسنامه و پرونده جامع قطعه")
        existing_parts_list = pd.read_sql_query("SELECT part_code FROM production", conn)['part_code'].tolist()
        
        search_code = st.selectbox("انتخاب یا جستجوی کد قطعه جهت مشاهده شناسنامه کامل:", ["-- انتخاب کنید --"] + existing_parts_list)
        
        if search_code and search_code != "-- انتخاب کنید --":
            prod_df = pd.read_sql_query("SELECT * FROM production WHERE part_code=?", conn, params=(search_code,))
            qc_df = pd.read_sql_query("SELECT * FROM qc WHERE part_code=?", conn, params=(search_code,))
            cost_df = pd.read_sql_query("SELECT * FROM cost_calculator WHERE part_code=?", conn, params=(search_code,))
            
            if not prod_df.empty:
                p_row = prod_df.iloc[0]
                st.success(f"اطلاعات کامل پرونده قطعه {search_code} با تاریخ شمسی یافت شد.")
                tab1, tab2, tab3, tab4 = st.tabs(["📌 پارامترهای تولید", "🧪 اطلاعات پودر مصرفی", "🔬 کنترل کیفیت QC", "💰 برآورد مالی"])
                
                with tab1:
                    st.subheader("مشخصات فنی و فرآیند ساخت")
                    prod_fmt = format_production_df_view(prod_df)
                    clean_prod = prod_fmt.drop(columns=['checklist_json', 'qc_checks_json', 'finishing_json'], errors='ignore')
                    disp_prod = clean_prod.rename(columns=FARSI_HEADERS_MAP)
                    st.table(disp_prod.T)
                    
                powder_code_val = p_row['powder_code']
                powder_df = pd.read_sql_query("SELECT * FROM powders WHERE powder_code=?", conn, params=(powder_code_val,))
                powder_chk_dict = {}
                if powder_df.empty:
                    powder_df = pd.read_sql_query("SELECT * FROM nora_powders WHERE powder_code=?", conn, params=(powder_code_val,))
                else:
                    try:
                        powder_chk_dict = json.loads(powder_df.iloc[0].get('checklist_json', '{}'))
                    except Exception:
                        pass

                with tab2:
                    if not powder_df.empty:
                        st.write(f"**کد ظرف پودر استفاده شده:** `{powder_code_val}`")
                        flat_powder = flatten_powder_df(powder_df)
                        st.table(flat_powder.T)
                    else:
                        st.warning("اطلاعات پودر متناظر یافت نشد.")
                        
                qc_chk_dict = {}
                with tab3:
                    if not qc_df.empty:
                        flat_qc = flatten_qc_df(qc_df)
                        st.table(flat_qc.T)
                        try:
                            qc_chk_dict = json.loads(qc_df.iloc[0].get('qc_checks_json', '{}'))
                        except Exception:
                            pass
                    else:
                        st.info("فرم کنترل کیفیت برای این قطعه هنوز ثبت نشده است.")
                        
                with tab4:
                    if not cost_df.empty:
                        cost_fmt = format_cost_df_view(cost_df)
                        disp_cost = cost_fmt.rename(columns=FARSI_HEADERS_MAP)
                        st.table(disp_cost.T)
                        st.metric("قیمت نهایی فروش (ریال)", f"{cost_df['final_price'].values[0]:,.0f}")
                    else:
                        st.info("محاسبه هزینه برای این قطعه ثبت نشده است.")
                
                dossier_sections = {
                    "۱- مشخصات و شناسنامه فرآیند تولید قطعه": [
                        ("کد قطعه", search_code),
                        ("نام قطعه", p_row['part_name']),
                        ("مدل دستگاه", p_row['machine_model']),
                        ("شماره ظرف پودر", powder_code_val),
                        ("تعداد روی صفحه", p_row['quantity']),
                        ("تاریخ ساخت (شمسی)", format_jalali_date(p_row['date'])),
                        ("زمان تولید (ساعت:دقیقه)", hours_to_time_str(p_row['build_time_hrs'])),
                        ("زمان توقف حین ساخت", hours_to_time_str(p_row['downtime_hrs'])),
                        ("تاریخ و ساعت شروع", f"{format_jalali_date(p_row['start_date'])} {p_row['start_time']}"),
                        ("تاریخ و ساعت پایان", f"{format_jalali_date(p_row['end_date'])} {p_row['end_time']}"),
                        ("پودر ورودی به دستگاه (گرم)", f"{p_row['input_powder_g']:,.1f}"),
                        ("پودر غیرقابل بازیافت (گرم)", f"{p_row['waste_powder_g']:,.1f}"),
                        ("وزن قطعه با ساپورت (گرم)", f"{p_row['part_with_support_g']:,.1f}"),
                        ("وزن قطعه نهایی (گرم)", f"{p_row['final_part_g']:,.1f}"),
                        ("درصد فیلتر دستگاه", f"{p_row['filter_percentage']}%"),
                        ("کد صفحه ساخت", p_row['build_plate_code'] or "---")
                    ]
                }

                if not powder_df.empty:
                    pwd_row = powder_df.iloc[0]
                    dossier_sections["۲- مشخصات پودر مصرفی"] = [
                        ("کد ظرف پودر", pwd_row['powder_code']),
                        ("نوع متریال پودر", pwd_row.get('material', '---')),
                        ("وزن پودر (گرم)", f"{pwd_row.get('weight_g', 0):,.1f}"),
                        ("تاریخ ورود / تست", format_jalali_date(pwd_row.get('date', '')))
                    ]
                    if powder_chk_dict:
                        dossier_sections["چک‌لیست آزمون‌های خواص پودر مصرفی"] = powder_chk_dict

                if not qc_df.empty:
                    q_row = qc_df.iloc[0]
                    dossier_sections["۳- اطلاعات و مسئولین کنترل کیفیت"] = [
                        ("بازرس کنترل کیفیت", q_row.get('qc_inspector', '---')),
                        ("مسئول فنی / مهندسی کیفیت", q_row.get('qc_engineer', '---')),
                        ("مدیر تضمین کیفیت", q_row.get('qa_manager', '---')),
                        ("تاریخ ثبت بازرسی", format_jalali_date(q_row.get('date', '')))
                    ]
                    if qc_chk_dict:
                        dossier_sections["نتایج تست‌های کیفی، آزمون‌ها و ملاحظات (QC)"] = qc_chk_dict

                if not cost_df.empty:
                    cst_row = cost_df.iloc[0]
                    dossier_sections["۴- برآورد مالی و قیمت نهایی"] = [
                        ("نوع پودر فلزی", cst_row['powder_type']),
                        ("حجم قطعه (cm3)", f"{cst_row['volume_cm3']:.2f}"),
                        ("وزن خالص قطعه (گرم)", f"{cst_row['net_weight_g']:.1f}"),
                        ("زمان کل چاپ", hours_to_time_str(cst_row['print_time_hrs'])),
                        ("هزینه پودر (ریال)", f"{cst_row['powder_cost_total']:,.0f}"),
                        ("استهلاک دستگاه (ریال)", f"{cst_row['depreciation_cost_total']:,.0f}"),
                        ("هزینه گاز و برق (ریال)", f"{cst_row['argon_cost_total'] + cst_row['power_cost_total']:,.0f}"),
                        ("دستمزد مهندسی و اپراتور (ریال)", f"{cst_row['engineering_cost_total'] + cst_row['operator_cost_total']:,.0f}"),
                        ("بهای تمام شده کل (ریال)", f"{cst_row['total_production_cost']:,.0f}"),
                        ("قیمت نهایی قابل ارائه به مشتری (ریال)", f"{cst_row['final_price']:,.0f}")
                    ]

                excel_dossier_bytes = build_form_layout_excel(dossier_sections, title=f"شناسنامه پرونده جامع قطعه {search_code}")
                st.download_button(
                    label="📥 دانلود خروجی اکسل",
                    data=excel_dossier_bytes,
                    file_name=f"part_dossier_form_{search_code}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_down_part_full_styled"
                )

        st.markdown("---")
        st.subheader("📂 مشاهده و دانلود چندانتخابی جداول پایگاه داده")
        
        table_options_map = {
            "آنالیز پودر اولیه": "powders",
            "پودرهای بازیافت شده": "recycled_powders",
            "پودرهای خریداری شده از نورا": "nora_powders",
            "رکوردهای تولید": "production",
            "کنترل کیفیت QC": "qc",
            "محاسبه هزینه و مالی": "cost_calculator"
        }
        
        selected_table_labels = st.multiselect(
            "جداول مورد نظر جهت مشاهده و خروجی اکسل را انتخاب کنید (یک یا چند مورد):",
            list(table_options_map.keys()),
            default=list(table_options_map.keys())[:1]
        )
        
        if selected_table_labels:
            multisheet_export_dict = {}
            json_cols_to_drop = ['checklist_json', 'qc_checks_json', 'finishing_json']
            
            for tbl_label in selected_table_labels:
                tbl_name = table_options_map[tbl_label]
                raw_df = pd.read_sql_query(f"SELECT * FROM {tbl_name}", conn)
                
                if tbl_name == "powders":
                    farsi_df = flatten_powder_df(raw_df)
                elif tbl_name == "qc":
                    farsi_df = flatten_qc_df(raw_df)
                elif tbl_name == "production":
                    prod_fmt = format_production_df_view(raw_df)
                    clean_df = prod_fmt.drop(columns=[col for col in json_cols_to_drop if col in prod_fmt.columns], errors='ignore')
                    farsi_df = clean_df.rename(columns=FARSI_HEADERS_MAP)
                elif tbl_name == "cost_calculator":
                    cost_fmt = format_cost_df_view(raw_df)
                    clean_df = cost_fmt.drop(columns=[col for col in json_cols_to_drop if col in cost_fmt.columns], errors='ignore')
                    farsi_df = clean_df.rename(columns=FARSI_HEADERS_MAP)
                elif tbl_name in ["recycled_powders", "nora_powders"]:
                    raw_copy = raw_df.copy()
                    if 'date' in raw_copy.columns:
                        raw_copy['date'] = raw_copy['date'].apply(format_jalali_date)
                    clean_df = raw_copy.drop(columns=[col for col in json_cols_to_drop if col in raw_copy.columns], errors='ignore')
                    farsi_df = clean_df.rename(columns=FARSI_HEADERS_MAP)
                else:
                    clean_df = raw_df.drop(columns=[col for col in json_cols_to_drop if col in raw_df.columns], errors='ignore')
                    farsi_df = clean_df.rename(columns=FARSI_HEADERS_MAP)
                
                st.markdown(f"#### 📊 {tbl_label}")
                if not farsi_df.empty:
                    st.table(farsi_df)
                else:
                    st.info(f"جدول {tbl_label} در حال حاضر خالی است.")
                
                multisheet_export_dict[tbl_label] = farsi_df

            st.markdown("---")
            combined_excel_file = export_to_styled_excel_multisheet(multisheet_export_dict, "archive_export.xlsx")
            st.download_button(
                label="📥 دانلود خروجی اکسل",
                data=combined_excel_file,
                file_name="selected_tables_archive.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_down_multiselect_archive"
            )

        conn.close()

    # ---------------------------------------------------------
    # ۷. مدیریت کامل کاربران
    # ---------------------------------------------------------
    elif choice == "👥 مدیریت کاربران":
        st.header("👥 مدیریت کامل کاربران و سطوح دسترسی")
        
        conn = get_db_connection()
        users_df = pd.read_sql_query("SELECT username, full_name, role FROM users", conn)
        
        st.subheader("📋 لیست کاربران فعلی سامانه")
        st.table(users_df.rename(columns={'username': 'نام کاربری', 'full_name': 'نام و نام خانوادگی', 'role': 'سطح دسترسی'}))
        
        st.markdown("---")
        
        u_tab1, u_tab2, u_tab3 = st.tabs(["➕ ۱. تعریف کاربر جدید", "✏️ ۲. ویرایش کاربر موجود", "🗑️ ۳. حذف کاربر"])
        
        with u_tab1:
            with st.form("add_user_form"):
                u_col1, u_col2 = st.columns(2)
                with u_col1:
                    new_username = st.text_input("نام کاربری جدید (لاتین)")
                    new_password = st.text_input("رمز عبور", type="password")
                with u_col2:
                    new_fullname = st.text_input("نام و نام خانوادگی")
                    new_role = st.selectbox("سطح دسترسی / نقش", ["مدیریت", "اپراتور و طراح", "کنترل کیفیت", "بازرگانی"])
                
                submit_new_user = st.form_submit_button("💾 ثبت کاربر جدید")
                if submit_new_user:
                    if new_username and new_password:
                        c = conn.cursor()
                        c.execute("INSERT OR REPLACE INTO users (username, password_hash, full_name, role) VALUES (?, ?, ?, ?)",
                                  (new_username, hash_password(new_password), new_fullname, new_role))
                        conn.commit()
                        st.success(f"کاربر {new_username} با موفقیت تعریف شد.")
                        st.rerun()
                    else:
                        st.error("نام کاربری و رمز عبور الزامی است.")

        with u_tab2:
            all_usernames = users_df['username'].tolist()
            if all_usernames:
                selected_edit_user = st.selectbox("انتخاب کاربر جهت ویرایش:", all_usernames)
                user_info = pd.read_sql_query("SELECT * FROM users WHERE username=?", conn, params=(selected_edit_user,)).iloc[0]
                
                with st.form("edit_user_form"):
                    e_col1, e_col2 = st.columns(2)
                    with e_col1:
                        st.text_input("نام کاربری", value=user_info['username'], disabled=True)
                        edit_password = st.text_input("رمز عبور جدید (در صورت عدم تغییر خالی بگذارید)", type="password")
                    with e_col2:
                        edit_fullname = st.text_input("نام و نام خانوادگی", value=user_info['full_name'])
                        roles_list = ["مدیریت", "اپراتور و طراح", "کنترل کیفیت", "بازرگانی"]
                        default_role_idx = roles_list.index(user_info['role']) if user_info['role'] in roles_list else 0
                        edit_role = st.selectbox("سطح دسترسی / نقش", roles_list, index=default_role_idx)
                    
                    submit_edit_user = st.form_submit_button("💾 ذخیره تغییرات کاربر")
                    if submit_edit_user:
                        c = conn.cursor()
                        if edit_password:
                            c.execute("UPDATE users SET full_name=?, role=?, password_hash=? WHERE username=?",
                                      (edit_fullname, edit_role, hash_password(edit_password), selected_edit_user))
                        else:
                            c.execute("UPDATE users SET full_name=?, role=? WHERE username=?",
                                      (edit_fullname, edit_role, selected_edit_user))
                        conn.commit()
                        st.success(f"اطلاعات کاربر {selected_edit_user} با موفقیت به‌روزرسانی شد.")
                        st.rerun()

        with u_tab3:
            all_usernames = users_df['username'].tolist()
            deletable_users = [u for u in all_usernames if u != st.session_state["username"]]
            
            if deletable_users:
                selected_delete_user = st.selectbox("انتخاب کاربر جهت حذف از سیستم:", deletable_users)
                if st.button("⚠️ حذف کاربر انتخاب شده"):
                    c = conn.cursor()
                    c.execute("DELETE FROM users WHERE username=?", (selected_delete_user,))
                    conn.commit()
                    st.success(f"کاربر {selected_delete_user} با موفقیت حذف شد.")
                    st.rerun()
            else:
                st.info("کاربر دیگری جهت حذف وجود ندارد.")

        conn.close()
